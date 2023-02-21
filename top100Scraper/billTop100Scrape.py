from bs4 import BeautifulSoup
import requests
import re
import openpyxl
from openpyxl import Workbook, load_workbook
import json
import datetime
from scrapeID import spotifyID, spotifyToken
from openpyxl.utils import get_column_letter
currentTime = datetime.datetime.now()

class scrapeAndMake:
    def __init__(self):
        self.spotifyID = spotifyID
        self.spotifyToken = spotifyToken
        self.songInfo = {}

    def scrape(self):
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = 'Billboard Top 100'
        sheet.append(["{}/{}/{}".format(currentTime.day,currentTime.month, currentTime.year),'Song','Artist'])

        try:
            source = requests.get("https://www.billboard.com/charts/hot-100/")
            source.raise_for_status()
            mySoup = BeautifulSoup(source.text, 'html.parser')
            songs = mySoup.find('div', class_="chart-results-list // lrv-u-padding-t-150 lrv-u-padding-t-050@mobile-max").find_all(class_=re.compile("o-chart-results-list-row-container"))
            songRank = 1
            for song in songs:
                songName = song.find('h3', "c-title").get_text(strip=True)
                songArtist = song.find('span',class_=re.compile("c-label a-no-trucate")).get_text(strip=True)      
                sheet.append([songRank, songName, songArtist])
                firstName = re.split(" Featuring | & | X |, |,| x | with | With ", songArtist)
                self.songInfo["{}".format(songName)] = {"songName":songName, "artist":songArtist, "uri":self.getSongs(songName,firstName[0])}
                songRank = songRank + 1
        except Exception as e:
           print("The link is not valid!")
        excel.save("Billboard Top 100.xlsx")

    def makePlaylist(self):
                request_body = json.dumps({
                "name": "Billboard Top 100: {}/{}/{}".format(currentTime.day,currentTime.month, currentTime.year),
                "description": "The Billboard Top 100 for the specified time range! In descending order.",
                "public": True})

                query = "https://api.spotify.com/v1/users/{}/playlists".format(
                spotifyID)
                response = requests.post(
                query,
                data=request_body,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotifyToken)
                }
                )
                response_json = response.json()
                return response_json["id"]


    def getSongs(self, songName, artist):
            query = "https://api.spotify.com/v1/search?q=artist:{}%20track:{}&type=track&offset=0&limit=1".format(
                artist,
                songName
            )
            response = requests.get(
                query,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotifyToken)
                }
            )
            response_json = response.json()
            songs = response_json["tracks"]["items"]
            songNameAppended = songName
            if(len(songs) !=0):
                uri = songs[0]["uri"]
                if(uri is not None):
                    return uri
            if(len(songNameAppended) == 0):
                print("The URI could not be found.")
            else:
                songNameAppended = songNameAppended[:-1]
                return self.getSongs(songNameAppended, artist)

    def addSongs(self):
        playlist = self.makePlaylist()
        self.scrape()
       
        uri = [] 
        for song, info in self.songInfo.items():
            if(info["uri"] is not None):
                uri.append(info["uri"])

        data = json.dumps(uri)

        query = "https://api.spotify.com/v1/playlists/{}/tracks".format(
                playlist)

        response = requests.post(
                query,
                data=data,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotifyToken)
                }
            )

        self.songInfo.clear()
        response_json = response.json()
        return response_json

    def makePlaylistFromFile(self, fileWb):
                request_body = json.dumps({
                "name": "Billboard Top 100: {}".format(fileWb['A1'].value),
                "description": "The Billboard Top 100 for the specified time range! In descending order.",
                "public": True})

                query = "https://api.spotify.com/v1/users/{}/playlists".format(
                spotifyID)
                response = requests.post(
                query,
                data=request_body,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotifyToken)
                }
                )
                response_json = response.json()
                return response_json["id"]
        
    def addSongsFromFile(self, filePath):
        localWb = load_workbook(filePath)
        fileWb = localWb.active
        playlist = self.makePlaylistFromFile(fileWb)
        col1 = 2
        colLet1 = get_column_letter(col1)
        col2 = 3
        colLet2 = get_column_letter(col2)
        for row in range(2,102):
            songName = str(fileWb[colLet1 + str(row)].value)
            songArtist = str(fileWb[colLet2 + str(row)].value) 
            firstName = re.split(" Featuring | & | X |, |,| x | with | With ", songArtist)
            self.songInfo["{}".format(songName)] = {"songName":songName, "artist":songArtist, "uri":self.getSongs(songName,firstName[0])}
        
        uri = [] 
        for song, info in self.songInfo.items():
            if(info["uri"] is not None):
                uri.append(info["uri"])

        data = json.dumps(uri)

        query = "https://api.spotify.com/v1/playlists/{}/tracks".format(
                playlist)

        response = requests.post(
                query,
                data=data,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotifyToken)
                }
            )

        response_json = response.json()
        return response_json


if __name__ == "__main__":
    test = scrapeAndMake()
    test.addSongsFromFile("/Users/rohitnair/billTop100Scrape/top100Scraper/Billboard Top 100.xlsx")

        


